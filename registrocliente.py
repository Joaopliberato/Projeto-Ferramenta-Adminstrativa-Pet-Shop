import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import os
from PIL import ImageTk, Image

def register_customer():
    nome_cliente = customer_name_entry.get()
    endereco_cliente = customer_address_entry.get()
    telefone_cliente = customer_phone_entry.get()
    nome_pet = pet_name_entry.get()

    if not all([nome_cliente, endereco_cliente, telefone_cliente, nome_pet]):
        messagebox.showerror("Erro", "Por favor, preencha todos os campos!")
        return

    message = f"Cliente registrado com sucesso!\nNome: {nome_cliente}\nEndereço: {endereco_cliente}\nTelefone: {telefone_cliente}\nNome do Pet: {nome_pet}"
    messagebox.showinfo("Sucesso", message)

    clear_input_fields()
    save_data(nome_cliente, endereco_cliente, telefone_cliente, nome_pet)

def clear_input_fields():
    customer_name_entry.delete(0, tk.END)
    customer_address_entry.delete(0, tk.END)
    customer_phone_entry.delete(0, tk.END)
    pet_name_entry.delete(0, tk.END)

def save_data(nome_cliente, endereco_cliente, telefone_cliente, nome_pet):
    try:
        try:
            workbook = openpyxl.load_workbook("petshopclientes_data.xlsx")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        if sheet.max_row == 1:
            sheet['A1'] = "Nome do Cliente"
            sheet['B1'] = "Endereço do Cliente"
            sheet['C1'] = "Telefone do Cliente"
            sheet['D1'] = "Nome do Pet"

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = nome_cliente
        sheet.cell(row=next_row, column=2).value = endereco_cliente
        sheet.cell(row=next_row, column=3).value = telefone_cliente
        sheet.cell(row=next_row, column=4).value = nome_pet

        # Estilizar as células
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        workbook.save("petshopclientes_data.xlsx")
        messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

def go_back():
    janela.destroy()
    os.system('python telaini.py')

# Criar janela principal
janela = tk.Tk()
janela.title("Ferramenta Administrativa - Aroa Pet Shop")
janela.geometry("800x600")
janela.configure(background="black")

# Carregar a imagem de fundo
background_image = Image.open("pic/degrade.jpg")
background_image = background_image.resize((800, 600))
background_image_tk = ImageTk.PhotoImage(background_image)

# Criar um canvas para exibir a imagem de fundo
canvas = tk.Canvas(janela, width=800, height=600)
canvas.pack(fill="both", expand=True)

# Exibir a imagem de fundo no canvas
canvas.create_image(0, 0, image=background_image_tk, anchor="nw")

# Estilos
style = ttk.Style()
style.theme_use("default")

# Configurar o estilo dos rótulos
style.configure("Black.TLabel",
                font=("Montserrat", 12, "bold"),
                foreground="#000000")

# Configurar o estilo dos botões
style.configure("Black.TButton",
                font=("Montserrat", 12, "bold"),
                foreground="#000000")

content_frame = ttk.Frame(canvas, padding=(20, 20))
content_frame.place(relx=0.5, rely=0.5, anchor="center")

# Rótulo e campo de nome do cliente
customer_name_label = ttk.Label(content_frame, text="Nome:", style="Black.TLabel")
customer_name_label.grid(row=0, column=0, sticky=tk.E)
customer_name_entry = ttk.Entry(content_frame, width=30)
customer_name_entry.grid(row=0, column=1, padx=10, pady=5)

# Rótulo e campo de endereço do cliente
customer_address_label = ttk.Label(content_frame, text="Endereço:", style="Black.TLabel")
customer_address_label.grid(row=1, column=0, sticky=tk.E)
customer_address_entry = ttk.Entry(content_frame, width=30)
customer_address_entry.grid(row=1, column=1, padx=10, pady=5)

# Rótulo e campo de telefone do cliente
customer_phone_label = ttk.Label(content_frame, text="Telefone:", style="Black.TLabel")
customer_phone_label.grid(row=2, column=0, sticky=tk.E)
customer_phone_entry = ttk.Entry(content_frame, width=30)
customer_phone_entry.grid(row=2, column=1, padx=10, pady=5)

# Rótulo e campo de nome do pet
pet_name_label = ttk.Label(content_frame, text="Nome do Pet:", style="Black.TLabel")
pet_name_label.grid(row=3, column=0, sticky=tk.E)
pet_name_entry = ttk.Entry(content_frame, width=30)
pet_name_entry.grid(row=3, column=1, padx=10, pady=5)

# Botão de registro de cliente
register_button = ttk.Button(content_frame, text="Registrar", command=register_customer, style="Black.TButton")
register_button.grid(row=4, column=1, pady=10)

# Botão Voltar
back_button = ttk.Button(content_frame, text="Voltar", command=go_back, style="Black.TButton")
back_button.grid(row=4, column=0, pady=10)

# Configuração das colunas
content_frame.columnconfigure(1, weight=1)

janela.mainloop()