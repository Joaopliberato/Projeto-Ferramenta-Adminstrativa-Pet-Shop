import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import os
from PIL import ImageTk, Image
from openpyxl.styles import Font, Alignment

def register_service():
    tipo_servico = service_type_combobox.get()
    nome_cliente = customer_name_entry.get()
    numero_cliente = customer_phone_entry.get()
    pontuacao = customer_score_entry.get()

    if not all([tipo_servico, nome_cliente, numero_cliente, pontuacao]):
        messagebox.showerror("Erro", "Por favor, preencha todos os campos!")
        return

    price = service_prices[tipo_servico]

    message = f"Serviço registrado com sucesso!.\nTipo de serviço: {tipo_servico}\nPreço: {price}\nCliente: {nome_cliente}\nPontuação: {pontuacao}"
    messagebox.showinfo("Sucesso", message)

    clear_input_fields()
    save_data(tipo_servico, price, nome_cliente, numero_cliente, pontuacao)

def clear_input_fields():
    service_type_combobox.set('')
    customer_name_entry.delete(0, tk.END)
    customer_phone_entry.delete(0, tk.END)
    customer_score_entry.delete(0, tk.END)

def save_data(service_type, price, customer_name, customer_phone, customer_score):
    try:
        try:
            workbook = openpyxl.load_workbook("petshop_data.xlsx")
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # Adicionar nomes às colunas
        if sheet.max_row == 1:
            sheet.cell(row=1, column=1).value = "Tipo de Serviço"
            sheet.cell(row=1, column=2).value = "Preço"
            sheet.cell(row=1, column=3).value = "Nome do Cliente"
            sheet.cell(row=1, column=4).value = "Número do Cliente"
            sheet.cell(row=1, column=5).value = "Pontuação"

        next_row = sheet.max_row + 1
        sheet.cell(row=next_row, column=1).value = service_type
        sheet.cell(row=next_row, column=2).value = price
        sheet.cell(row=next_row, column=3).value = customer_name
        sheet.cell(row=next_row, column=4).value = customer_phone
        sheet.cell(row=next_row, column=5).value = customer_score

        # Aplicar formatação às células
        font = Font(bold=True)
        alignment = Alignment(horizontal="center")
        for column in range(1, 6):
            cell = sheet.cell(row=1, column=column)
            cell.font = font
            cell.alignment = alignment

            cell = sheet.cell(row=next_row, column=column)
            cell.font = font
            cell.alignment = alignment

        workbook.save("petshop_data.xlsx")
        messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

def go_back():
    janela.destroy()
    os.system('python telaini.py')

service_prices = {
    "Banho": "R$ 35,00",
    "Tosa": "R$ 25,00",
    "Banho e Tosa": "R$ 50,00",
    "Vermifugação": "R$ 60,00"
}

# Criar janela principal
janela = tk.Tk()
janela.title("Ferramenta Administrativa - Aroa Pet Shop")
janela.geometry("800x600")
janela.configure(background="black")

# Estilos
style = ttk.Style()
style.theme_use("default")

# Configurar o estilo dos rótulos
style.configure("Black.TLabel",
                font=("Montserrat", 12, "bold"),
                foreground="black")

# Configurar o estilo dos botões
style.configure("Black.TButton",
                font=("Montserrat", 12, "bold"),
                foreground="black")

# Carregar a imagem de fundo
background_image = Image.open("pic/degrade.jpg")
background_image = background_image.resize((800, 600))
background_image_tk = ImageTk.PhotoImage(background_image)

# Criar um canvas para exibir a imagem de fundo
canvas = tk.Canvas(janela, width=800, height=600)
canvas.pack(fill="both", expand=True)

# Exibir a imagem de fundo no canvas
canvas.create_image(0, 0, image=background_image_tk, anchor="nw")

content_frame = ttk.Frame(canvas, padding=(20, 20))
content_frame.place(relx=0.5, rely=0.5, anchor="center")

# Tipos de serviço
service_types = list(service_prices.keys())

# Combobox para seleção do tipo de serviço
service_type_combobox = ttk.Combobox(content_frame, values=service_types, state="readonly")
service_type_combobox.grid(row=0, column=1, padx=10, pady=5)

# Rótulo e campo de preço
price_label = ttk.Label(content_frame, text="Preço:", style="Black.TLabel")
price_label.grid(row=1, column=0, sticky=tk.E)
price_value = tk.StringVar()
price_value.set(service_prices[service_types[0]])
price_entry = ttk.Entry(content_frame, width=30, textvariable=price_value, state="readonly")
price_entry.grid(row=1, column=1, padx=10, pady=5)

# Rótulo e campo de nome do cliente
customer_name_label = ttk.Label(content_frame, text="Nome:", style="Black.TLabel")
customer_name_label.grid(row=2, column=0, sticky=tk.E)
customer_name_entry = ttk.Entry(content_frame, width=30)
customer_name_entry.grid(row=2, column=1, padx=10, pady=5)

# Rótulo e campo de número do cliente
customer_phone_label = ttk.Label(content_frame, text="Número:", style="Black.TLabel")
customer_phone_label.grid(row=3, column=0, sticky=tk.E)
customer_phone_entry = ttk.Entry(content_frame, width=30)
customer_phone_entry.grid(row=3, column=1, padx=10, pady=5)

# Rótulo e campo de pontuação do cliente
customer_score_label = ttk.Label(content_frame, text="Pontuação:", style="Black.TLabel")
customer_score_label.grid(row=4, column=0, sticky=tk.E)
customer_score_entry = ttk.Entry(content_frame, width=30)
customer_score_entry.grid(row=4, column=1, padx=10, pady=5)

# Botão de registro de serviço
register_button = ttk.Button(content_frame, text="Registrar", command=register_service, style="Black.TButton")
register_button.grid(row=5, column=1, pady=10)

# Botão Voltar
back_button = ttk.Button(content_frame, text="Voltar", command=go_back, style="Black.TButton")
back_button.grid(row=5, column=0, pady=10)

# Função para atualizar o preço quando um tipo de serviço é selecionado
def update_price(*args):
    service_type = service_type_combobox.get()
    if service_type in service_prices:
        price_value.set(service_prices[service_type])

# Vincula a função de atualização ao evento de seleção do combobox
service_type_combobox.bind("<<ComboboxSelected>>", update_price)

# Configuração das colunas
content_frame.columnconfigure(1, weight=1)

janela.mainloop()